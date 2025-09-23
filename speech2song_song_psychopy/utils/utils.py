import os
import yaml
import openpyxl
from datetime import datetime
from psychopy import core, event
from psychopy import sound
import time

def save_params_to_yaml(params, filename="run_params.yaml"):
    """
    Saves parameters to a YAML file. If the file already exists, 
    appends the new parameters to it.
    
    Args:
        params (dict): The parameters to save, stored as a dictionary.
        filename (str): The YAML file name to save to. Defaults to "run_params.yaml".
    """
    # Get the current timestamp to record when the run happened
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    params_with_timestamp = {"timestamp": timestamp, "params": params}
    
    # Append the parameters to the YAML file
    with open(filename, 'a') as file:
        yaml.dump([params_with_timestamp], file, default_flow_style=False)

def write_value_to_excel(file_path, sheet_name, value, row_name, column_name):
    ''' Write a vector to specified cell in an Excel sheet '''
    print(f"Writing to {file_path} on sheet {sheet_name}...")
    if os.path.exists(file_path):
        # If file exists, open it
        workbook = openpyxl.load_workbook(file_path)
        print("Workbook opened.")
        # If sheet doesn't exist, create it
        if sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            print(f"Sheet {sheet_name} found.")
        else:
            sheet = workbook.create_sheet(sheet_name)
            print(f"Sheet {sheet_name} created.")
    else:
        # If file doesn't exist, create it
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = sheet_name
        print("Workbook created with new sheet.")
    # find the row and column
    for row in range(1, sheet.max_row + 1):
        for column in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(row=row, column=column).value
            if cell_value == row_name:
                write_row = row
    for row in range(1, sheet.max_row + 1):
        for column in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(row=row, column=column).value
            if cell_value == column_name:
                write_column = column
    # Write the vector into specified cell
    sheet.cell(row=write_row, column=write_column, value=value)
    
    # Save the workbook
    workbook.save(file_path)
    print(f"Data saved to {file_path}.")

def write_vector_to_excel_col(file_path, sheet_name, vector):
    ''' write a vector as a col into excel '''
    print(f"Writing to {file_path} on sheet {sheet_name}...")
    if os.path.exists(file_path):
        # if file exists, open it
        workbook = openpyxl.load_workbook(file_path)
        print("Workbook opened.")
        print(workbook.sheetnames)
        # if sheet doesn't exist, create
        if sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            print(f"Sheet {sheet_name} found.")
        else:
            sheet = workbook.create_sheet(sheet_name)
            print(f"Sheet {sheet_name} created.")
    else:
        # if file doesn't exist, create
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = sheet_name
        sheet.delete_rows(1, 1)
        print("Workbook created with new sheet.")

    # find the next row (considering existing data)
    if sheet.max_row == 1:
        next_col = 1
    else:
        next_col = sheet.max_column + 1
    print(f"Next available column: {next_col}")

    # write the vector into next row
    for row_num, value in enumerate(vector, start=1):
        sheet.cell(row=row_num, column=next_col, value=value)
    
    # save the workbook
    workbook.save(file_path)
    print(f"Data saved to {file_path}.")

def wait_for_mouse_click(mouse):
    core.wait(0.5)
    mouse.clickReset()
    Pressed = False
    while Pressed == False:
        if any(mouse.getPressed()):
            Pressed = True

def wait_for_enter_key(end_key_name):
    core.wait(0.5)
    event.clearEvents(eventType='keyboard')
    pressed = False
    while not pressed:
        keys = event.getKeys()
        if end_key_name in keys:
            pressed = True

def getRelativeMousePos(mouse, win):
    mousePos = mouse.getPos()
    relativeMousePos = (
        (mousePos[0] / win.size[0]) * 2 - 1,  # x
        (mousePos[1] / win.size[1]) * 2 - 1   # y
    )
    return relativeMousePos

class Clock:
    def __init__(self):
        self.start_time = time.time()
    def getTime(self):
        return time.time() - self.start_time

def play_sound(file, volume):
	mySound = sound.Sound(file, preBuffer=-1, volume=volume)	
	mySound.play()
	core.wait(mySound.getDuration()+1)
     
def play_sound_rep(file, volume, n_repeat, gap, wait_time):
    time.sleep(wait_time)
    for r in range(n_repeat):
        mySound = sound.Sound(file, volume=volume)	
        repClock = Clock()
        cur_sound_start = repClock.getTime()
        mySound.play()
        while True:
            if repClock.getTime() > cur_sound_start + mySound.getDuration() + gap:
                break
            time.sleep(0.1)