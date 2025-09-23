import os
import yaml
import openpyxl
from datetime import datetime
from psychopy import core, event
from psychopy import sound
import time

def save_params_to_yaml(params, filename="run_params.yaml"):
    """
    Saves parameters to a YAML file. If the file already exists, 
    appends the new parameters to it.
    
    Args:
        params (dict): The parameters to save, stored as a dictionary.
        filename (str): The YAML file name to save to. Defaults to "run_params.yaml".
    """
    # Get the current timestamp to record when the run happened
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    params_with_timestamp = {"timestamp": timestamp, "params": params}
    
    # Append the parameters to the YAML file
    with open(filename, 'a') as file:
        yaml.dump([params_with_timestamp], file, default_flow_style=False)

def write_value_to_excel(file_path, sheet_name, value, row_key, column_name, row_key_column='sound'):
    '''
    Writes `value` to the Excel sheet at the cell where:
      - the row is determined by matching `row_key` in the column `row_key_column`
      - the column is determined by `column_name` in the first row
    '''
    print(f"Writing to {file_path} on sheet '{sheet_name}'...")

    if os.path.exists(file_path):
        workbook = openpyxl.load_workbook(file_path)
        if sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
        else:
            raise ValueError(f"Sheet '{sheet_name}' not found in {file_path}")
    else:
        raise FileNotFoundError(f"File '{file_path}' not found")

    # === Find the header row (usually row 1)
    headers = [sheet.cell(row=1, column=col).value for col in range(1, sheet.max_column + 1)]

    # Find column index of the column_name (e.g., "duration")
    if column_name not in headers:
        raise ValueError(f"Column '{column_name}' not found in sheet '{sheet_name}'.")
    write_column = headers.index(column_name) + 1

    # Find column index of the row key column (e.g., "sound")
    if row_key_column not in headers:
        raise ValueError(f"Row key column '{row_key_column}' not found in sheet '{sheet_name}'.")
    row_key_col_idx = headers.index(row_key_column) + 1

    # Find the row index where row_key matches
    write_row = None
    for row in range(2, sheet.max_row + 1):
        if str(sheet.cell(row=row, column=row_key_col_idx).value) == str(row_key):
            write_row = row
            break

    if write_row is None:
        raise ValueError(f"Row with {row_key_column}='{row_key}' not found in sheet '{sheet_name}'.")

    # Write the value
    sheet.cell(row=write_row, column=write_column, value=value)
    workbook.save(file_path)
    print(f"[OK] Wrote value to row {write_row}, column {write_column}.")
    
def write_vector_to_excel_col(file_path, sheet_name, vector):
    ''' write a vector as a col into excel '''
    print(f"Writing to {file_path} on sheet {sheet_name}...")
    if os.path.exists(file_path):
        # if file exists, open it
        workbook = openpyxl.load_workbook(file_path)
        print("Workbook opened.")
        print(workbook.sheetnames)
        # if sheet doesn't exist, create
        if sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            print(f"Sheet {sheet_name} found.")
        else:
            sheet = workbook.create_sheet(sheet_name)
            print(f"Sheet {sheet_name} created.")
    else:
        # if file doesn't exist, create
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = sheet_name
        sheet.delete_rows(1, 1)
        print("Workbook created with new sheet.")

    # find the next row (considering existing data)
    if sheet.max_row == 1:
        next_col = 1
    else:
        next_col = sheet.max_column + 1
    print(f"Next available column: {next_col}")

    # write the vector into next row
    for row_num, value in enumerate(vector, start=1):
        sheet.cell(row=row_num, column=next_col, value=value)
    
    # save the workbook
    workbook.save(file_path)
    print(f"Data saved to {file_path}.")

def wait_for_mouse_click(mouse):
    core.wait(0.5)
    mouse.clickReset()
    Pressed = False
    while Pressed == False:
        if any(mouse.getPressed()):
            Pressed = True

def wait_for_enter_key(end_key_name):
    core.wait(0.5)
    event.clearEvents(eventType='keyboard')
    pressed = False
    while not pressed:
        keys = event.getKeys()
        if end_key_name in keys:
            pressed = True

def getRelativeMousePos(mouse, win):
    mousePos = mouse.getPos()
    relativeMousePos = (
        (mousePos[0] / win.size[0]) * 2 - 1,  # x
        (mousePos[1] / win.size[1]) * 2 - 1   # y
    )
    return relativeMousePos

class Clock:
    def __init__(self):
        self.start_time = time.time()
    def getTime(self):
        return time.time() - self.start_time

def play_sound(file, volume):
	mySound = sound.Sound(file, preBuffer=-1, volume=volume)	
	mySound.play()
	core.wait(mySound.getDuration()+1)
     
def play_sound_rep(file, volume, n_repeat, gap, wait_time):
    time.sleep(wait_time)
    for r in range(n_repeat):
        mySound = sound.Sound(file, volume=volume)	
        repClock = Clock()
        cur_sound_start = repClock.getTime()
        mySound.play()
        while True:
            if repClock.getTime() > cur_sound_start + mySound.getDuration() + gap:
                break
            time.sleep(0.1)