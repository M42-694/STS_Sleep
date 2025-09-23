# --- Import packages ---
from psychopy import prefs
from psychopy import plugins
from psychopy import sound, gui, visual, core, data, event, logging, clock, colors, layout, hardware
from psychopy.constants import (NOT_STARTED, STARTED, PLAYING, PAUSED,
                                STOPPED, FINISHED, PRESSED, RELEASED, FOREVER, priority)
from psychopy.event import Mouse
from psychopy.hardware import keyboard
from psychopy.visual.slider import Slider

from numpy import (sin, cos, tan, log, log10, pi, average,
                   sqrt, std, deg2rad, rad2deg, linspace, asarray)
from numpy.random import random, randint, choice
import pandas as pd
import os  # handy system and path functions
import random
from IPython import get_ipython
import threading
import time
import openpyxl
from collections import Counter
import sys
import math
plugins.activatePlugins()
prefs.hardware['audioLib'] = ['sounddevice', 'PTB', 'pyo', 'pygame'] #'ptb'
prefs.hardware['audioLatencyMode'] = '0'
defaultKeyboard = keyboard.Keyboard(backend='psychotoolbox')

from slider_routine_discrete import slider_routine_discrete
from slider_routine_discrete_key import slider_routine_discrete_key
from slider_routine_continuous import slider_routine_continuous
from slider_routine_continuous_key import slider_routine_continuous_key

# import psychtoolbox.audio
# print(psychtoolbox.audio.get_devices())

################################
# --- Parameter to control --- #
################################

# trial - continuous slider: record time after the sound 
record_after = 2   # (sec)
# interval: sound presents after fixation
interval_cross_sound = 1   # (sec)
# sound volume
volume = 1
# repetition
n_repeat = 8        # repetition time of stimulus
gap_time = 0.5      # the intetval between repetition (sec)
# interval: sound presents after click
rep_interval_cross_sound = interval_cross_sound #(sec) 

# test mode
test_mode = True
test_n_stimulus = 10

# number of Break
n_break = 2

# control mode
control_mode = 'key'   # 'key' for keyboard, 'mouse' for mouse
continue_key = 'return'  # 'return' for Enter
                       # 'down' for ↓

if control_mode == 'key':
    end_key_name = continue_key
    control_text = 'Press ' + end_key_name
elif control_mode == 'mouse':
    control_text = 'Click'

##################################
# --- Setup global variables --- #
##################################
# get current path
thisDir = os.getcwd()
#thisDir = thisDir.replace('\\', '/')
# create a device manager to handle hardware (keyboards, mice, mirophones, speakers, etc.)
deviceManager = hardware.DeviceManager()
# info about the experiment session
psychopyVersion = '2024.1.4'
expName = 'SpeechToSong' 
# experiment info
expInfo = {
    'participant': f"{randint(0, 999999):06.0f}",
    'session': '001',
    'date|hid': data.getDateStr(),
    'expName|hid': expName,
    'psychopyVersion|hid': psychopyVersion,
} 


############################
# --- Window Parameter --- #
############################
# create window experiment
win_size = (1280, 720) # (800, 600) # [1920, 1080]
ratio_text_size = 40
ratio_text_width= 1
ratio_cross = 40
contrast = 1

if test_mode:
    full_screen = False
else :
    full_screen = True



def setupData(expInfo, dataDir=None):
    # remove dialog-specific syntax from expInfo
    for key, val in expInfo.copy().items():
        newKey, _ = data.utils.parsePipeSyntax(key)
        expInfo[newKey] = expInfo.pop(key)
    # data file name stem = absolute path + name; later add .psyexp, .csv, .log, etc
    if dataDir is None:
        dataDir = thisDir
    filename = u'data/%s/%s_%s_%s_%s' % (expInfo['participant'],expInfo['participant'], expInfo['session'], expName, expInfo['date'])
    # make sure filename is relative to dataDir
    if os.path.isabs(filename):
        dataDir = os.path.commonprefix([dataDir, filename])
        filename = os.path.relpath(filename, dataDir)
    
    # an ExperimentHandler isn't essential but helps with data saving
    thisExp = data.ExperimentHandler(
        name=expName, version='',
        extraInfo=expInfo, runtimeInfo=None,
        originPath=thisDir,
        savePickle=True, saveWideText=True,
        dataFileName=dataDir + os.sep + filename, sortColumns='time'
    )
    thisExp.setPriority('thisRow.t', priority.CRITICAL)
    thisExp.setPriority('expName', priority.LOW)
    # return experiment handler
    return thisExp

def showExpInfoDlg(expInfo):
    ''' participant ID input interface '''
    # show participant info dialog
    dlg = gui.DlgFromDict(
        dictionary=expInfo, sortKeys=False, title=expName, alwaysOnTop=True
    )
    if dlg.OK == False:
        core.quit()  # user pressed cancel
    # return expInfo
    return expInfo
# log participant ID
expInfo = showExpInfoDlg(expInfo=expInfo)
# test Info
# expInfo['participant'] = '000'
# expInfo['date'] = '0000-00-00_00h00.00.000'

thisExp = setupData(expInfo=expInfo)

############################
# --- Slider Parameter --- #
############################
# Slider Parameter
startClock = core.Clock()
slider_width=.5
slider_height=.1
slider_orientation=0
#Failing to set ticks or labels in code
slider_ticks=[1,2,3,4,5]
#slider_labels="Strongly Disagree,Disagree,Neither Agree nor Disagree,Agree,Strongly Agree"
slider_labels=['Strongly Speech-like','Speech-like','Neither Speech-like nor Song-like','Song-like','Strongly Song-like']
slider_granularity=.1
slider_decimals=1
sliding = 0
slideSpeed = 3
oldRating=-1
marker_colour_discrete='blue'
marker_colour_continuous='red'
marker_size=.1
sliderColour='LightGray'
# keyboard mode step size
step_size = 0.5
slideSpeed = 10*slider_granularity
# error code
error_slider_data_null = -99

##################################
# --- Prepare Sound Sequence --- #
##################################
# randomize sound sequence
types = ["song", "speech", "unlabelled"]                          # illusion - song, control - speech
stim_type = types[0]
folder_input = 'Stimulus_trigger/'

def StimulusSeqInit(stimulus_list):
    ''' randomize stimulus sequence '''
    ''' modern Fisher–Yates shuffle '''
    # org_stimulus_list = stimulus_list.copy()
    # org_song_idx = [i for i, value in enumerate(org_stimulus_list['type']) if value == 'song']
    # org_speech_idx = [i for i, value in enumerate(org_stimulus_list['type']) if value == 'speech']
    # random.shuffle(org_song_idx)
    # random.shuffle(org_speech_idx)
    # n_stimulus = len(stimulus_list)
    # counts = Counter(stimulus_list['type'])
    # n_song = counts['song']
    # n_speech = counts['speech']
    # song_val = 1
    # speech_val = 0
    # bin_order = [song_val] * n_song + [speech_val] * (n_stimulus - n_song)
    # for j in range(3):
    #     random.shuffle(bin_order)
    # song_idx = [i for i, value in enumerate(bin_order) if value == song_val]
    # speech_idx = [i for i, value in enumerate(bin_order) if value == speech_val]
    # ran_order = bin_order.copy()
    # for idx in range(n_song):
    #     ran_order[song_idx[idx]] = org_song_idx[idx]
    # for idx in range(n_song):
    #     ran_order[speech_idx[idx]] = org_speech_idx[idx]
    n_stimulus = len(stimulus_list)
    ran_order = list(tuple(range(n_stimulus)))
    random.shuffle(ran_order)     
    ran_stimulus_list = stimulus_list.reindex(index = ran_order)
    ran_stimulus_list.reset_index(drop=True, inplace=True)
    # # check the uniquement of output list
    # if not len(set(ran_stimulus_list['sound'])) == len(ran_stimulus_list['sound']):
    #     print("Sound list randomization error")
    #     sys.exit(1)

    return ran_stimulus_list

def write_value_to_excel(file_path, sheet_name, value, row_name, column_name):
    ''' Write a vector to specified cell in an Excel sheet '''
    print(f"Writing to {file_path} on sheet {sheet_name}...")
    if os.path.exists(file_path):
        # If file exists, open it
        workbook = openpyxl.load_workbook(file_path)
        print("Workbook opened.")
        # If sheet doesn't exist, create it
        if sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            print(f"Sheet {sheet_name} found.")
        else:
            sheet = workbook.create_sheet(sheet_name)
            print(f"Sheet {sheet_name} created.")
    else:
        # If file doesn't exist, create it
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = sheet_name
        print("Workbook created with new sheet.")
    # find the row and column
    for row in range(1, sheet.max_row + 1):
        for column in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(row=row, column=column).value
            if cell_value == row_name:
                write_row = row
    for row in range(1, sheet.max_row + 1):
        for column in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(row=row, column=column).value
            if cell_value == column_name:
                write_column = column
    # Write the vector into specified cell
    sheet.cell(row=write_row, column=write_column, value=value)
    
    # Save the workbook
    workbook.save(file_path)
    print(f"Data saved to {file_path}.")

def write_vector_to_excel(file_path, sheet_name, vector):
    ''' write a vector as a row into excel '''
    print(f"Writing to {file_path} on sheet {sheet_name}...")
    if os.path.exists(file_path):
        # if file exists, open it
        workbook = openpyxl.load_workbook(file_path)
        print("Workbook opened.")
        print(workbook.sheetnames)
        # if sheet doesn't exist, create
        if sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            print(f"Sheet {sheet_name} found.")
        else:
            sheet = workbook.create_sheet(sheet_name)
            print(f"Sheet {sheet_name} created.")
    else:
        # if file doesn't exist, create
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = sheet_name
        sheet.delete_rows(1, 1)
        print("Workbook created with new sheet.")

    # find the next row (considering existing data)
    next_row = sheet.max_row + 1
    print(f"Next available row: {next_row}")

    # write the vector into next row
    for col_num, value in enumerate(vector, start=1):
        sheet.cell(row=next_row, column=col_num, value=value)
    
    # save the workbook
    workbook.save(file_path)
    print(f"Data saved to {file_path}.")


# get sound file list
file_list = []
sound_type = []
for type in types:
    cur_file_list = os.listdir(os.path.join(thisDir,folder_input, type))
    file_list = file_list + cur_file_list
    sound_type = sound_type + [type] * len(cur_file_list)

sound_list = pd.DataFrame({
        'sound': file_list,
        'type': sound_type
    })
for shf in range(2):
    sound_list = StimulusSeqInit(sound_list)
n_stimulus = len(sound_list)

# add properties of sound
sound_list['duration'] = None
sound_list['pre-rating'] = None
sound_list['pre-rating_confirm'] = None
sound_list['post-rating'] = None
sound_list['post-rating_confirm'] = None

with pd.ExcelWriter(os.path.join(thisExp.dataFileName+'.xlsx')) as writer:
    sound_list.to_excel(writer, sheet_name='SoundList')

# write sound list into Excel
# save_to_excel(os.path.join(thisExp.dataFileName+'.xlsx'), sound_list, 'SoundList')
# workbook = openpyxl.load_workbook(os.path.join(thisExp.dataFileName+'.xlsx'))
# slider_sheetname = 'slider'
# workbook.create_sheet(title=slider_sheetname)
# workbook.save(os.path.join(thisExp.dataFileName+'.xlsx'))

def setupLogging(filename):
    # this outputs to the screen, not a file
    _loggingLevel = logging.getLevel(
        prefs.piloting['pilotLoggingLevel'])
    logging.console.setLevel(_loggingLevel)
    # save a log file for detail verbose info
    logFile = logging.LogFile(filename+'.log', level=_loggingLevel)
    
    return logFile

logFile = setupLogging(filename=thisExp.dataFileName)

# --- Define some variables which will change depending on pilot mode ---
# work out from system args whether we are running in pilot mode
PILOTING = core.setPilotModeFromArgs()
# start off with values from experiment settings
_fullScr = False
_winSize = (1024, 768)
_loggingLevel = logging.getLevel('warning')
# if in pilot mode, apply overrides according to preferences
if PILOTING:
    # force windowed mode
    if prefs.piloting['forceWindowed']:
        _fullScr = False
        # set window size
        _winSize = prefs.piloting['forcedWindowSize']
    # override logging level
    _loggingLevel = logging.getLevel(
        prefs.piloting['pilotLoggingLevel']
    )

def setupWindow(expInfo=None, win=None):
    if PILOTING:
        logging.debug('Fullscreen settings ignored as running in pilot mode.')
    
    if win is None:
        # if not given a window to setup, make one
        win = visual.Window(
            size=_winSize, fullscr=_fullScr, screen=0,
            winType='pyglet', allowStencil=False,
            monitor='testMonitor', color=[0,0,0], colorSpace='rgb',
            backgroundImage='', backgroundFit='none',
            blendMode='avg', useFBO=True,
            units='height', 
            checkTiming=False  # we're going to do this ourselves in a moment
        )
    else:
        # if we have a window, just set the attributes which are safe to set
        win.color = [0,0,0]
        win.colorSpace = 'rgb'
        win.backgroundImage = ''
        win.backgroundFit = 'none'
        win.units = 'height'
    if expInfo is not None:
        # get/measure frame rate if not already in expInfo
        if win._monitorFrameRate is None:
            win.getActualFrameRate(infoMsg='Attempting to measure frame rate of screen, please wait...')
        expInfo['frameRate'] = win._monitorFrameRate
    win.mouseVisible = False
    win.hideMessage()
    # show a visual indicator if we're in piloting mode
    if PILOTING and prefs.piloting['showPilotingIndicator']:
        win.showPilotingIndicator()
    
    return win

def pauseExperiment(thisExp, win=None, timers=[], playbackComponents=[]):
    # if we are not paused, do nothing
    if thisExp.status != PAUSED:
        return
    
    # pause any playback components
    for comp in playbackComponents:
        comp.pause()
    # prevent components from auto-drawing
    win.stashAutoDraw()
    # make sure we have a keyboard
    defaultKeyboard = deviceManager.getDevice('defaultKeyboard')
    if defaultKeyboard is None:
        defaultKeyboard = deviceManager.addKeyboard(
            deviceClass='keyboard',
            deviceName='defaultKeyboard',
            backend='ioHub',
        )
    # run a while loop while we wait to unpause
    while thisExp.status == PAUSED:
        # check for quit (typically the Esc key)
        if defaultKeyboard.getKeys(keyList=['escape']):
            endExperiment(thisExp, win=win)
        # flip the screen
        win.flip()
    # if stop was requested while paused, quit
    if thisExp.status == FINISHED:
        endExperiment(thisExp, win=win)
    # resume any playback components
    for comp in playbackComponents:
        comp.play()
    # restore auto-drawn components
    win.retrieveAutoDraw()
    # reset any timers
    for timer in timers:
        timer.reset()

def endExperiment(thisExp, win=None):
    """
    End this experiment, performing final shut down operations.
    
    This function does NOT close the window or end the Python process - use `quit` for this.
    
    """
    if win is not None:
        # remove autodraw from all current components
        win.clearAutoDraw()
        # Flip one final time so any remaining win.callOnFlip() 
        # and win.timeOnFlip() tasks get executed
        win.flip()
    # mark experiment handler as finished
    thisExp.status = FINISHED
    # shut down eyetracker, if there is one
    if deviceManager.getDevice('eyetracker') is not None:
        deviceManager.removeDevice('eyetracker')
    logging.flush()        

def saveData(thisExp):

    filename = thisExp.dataFileName
    # these shouldn't be strictly necessary (should auto-save)
    thisExp.saveAsWideText(filename + '.csv', delim='auto')
    thisExp.saveAsPickle(filename)

# Wait for mouse click event
def wait_for_mouse_click(mouse):
    core.wait(0.5)
    mouse.clickReset()
    Pressed = False
    while Pressed == False:
        if any(mouse.getPressed()):
            Pressed = True

def wait_for_enter_key(end_key_name):
    core.wait(0.5)
    event.clearEvents(eventType='keyboard')
    pressed = False
    while not pressed:
        keys = event.getKeys()
        if end_key_name in keys:
            pressed = True

def getRelativeMousePos(mouse, win):
    mousePos = mouse.getPos()
    relativeMousePos = (
        (mousePos[0] / win.size[0]) * 2 - 1,  # x
        (mousePos[1] / win.size[1]) * 2 - 1   # y
    )
    return relativeMousePos


######################
## --- main run
######################
#def run():
win = visual.Window(
    size=win_size,
    units="pix",
    fullscr=full_screen,
    color=[0, 0, 0])
# get screen refresh rate
# refresh_rate = win.getActualFrameRate()
# print("screen refresh rate:", refresh_rate, "Hz")

# fixation cross
fixation = visual.shape.ShapeStim(win=win, 
								  vertices="cross", 
								  color='white', 
								  fillColor='white', 
								  contrast=contrast, 
								  size=max(win.size)/ratio_cross)
# mouse will be used throughout the experiment
mouse = Mouse(visible=True, newPos=[0,0])
mouse = event.Mouse(win=win)
#-----------------------
# Welcome Instruction
#-----------------------
text = visual.TextStim(win=win, text="Welcome to our experiment!\
    \n\nDuring the experiment, you will listen to some speech excerpts and you will\
    \n be asked to rate how much each stimulus sounds as speech-like or song-like.\
    \n\n " + control_text +" to continue!",\
    color="white",
    contrast=contrast,
    wrapWidth=max(win.size)/ratio_text_width,
    height= max(win.size)/ratio_text_size)
text.autoDraw = True 
win.flip()
if control_mode == 'key':
    wait_for_enter_key(end_key_name)
elif control_mode == 'mouse':
    wait_for_mouse_click(mouse)
text.autoDraw = False    
win.flip()  
#-----------------------
# Trial Instruction
#-----------------------
n_sound = n_stimulus
text = visual.TextStim(win=win, text="The experiment consists of " + str(n_sound) + " sound excerpts as " + str(n_break+1) + " blocks. "\
	"\n\n For each excerpt, you will listen to it once at first, after which you will\
    \n be asked to rate how much it sounds as speech-like or song-like.\
	\n\n Then the same excerpt will be repeated " + str(n_repeat) +" times, during the repetition you can \
	\n move the mouse to rate the excerpt continously along the scale.\
    \n\n After the repetition you will listen to the same excerpt one more time, \
	\n and please give the final rate of the excerpt by clicking.\
	\n\n" + control_text + " to continue!",\
	color="white",
	contrast=contrast,
	wrapWidth=max(win.size)/ratio_text_width,
	height= max(win.size)/ratio_text_size)
text.autoDraw = True 
win.flip()
if control_mode == 'key':
    wait_for_enter_key(end_key_name)
elif control_mode == 'mouse':
    wait_for_mouse_click(mouse)
text.autoDraw = False    
win.flip()  

# Initialize components for Routine "trial"
(win_width, win_height) = win.size
trialClock = core.Clock()
slider_shape_width = 0.5
slider_shape_height = 0.5
margin_ratio_horizontal = 1.1
slider_shape = visual.Rect(win=win, name='slider_shape',
    width= win_width*slider_shape_width*margin_ratio_horizontal, height=win_height*slider_shape_height, 
    ori=0, pos=(0, 0), lineWidth=1, lineColor=[0,0,0], lineColorSpace='rgb',
    fillColor=[0,0,0], fillColorSpace='rgb', opacity=1, depth=-2.0, interpolate=True, )

slider = visual.Slider(win=win, name='slider',
    size=(win_width*slider_width, slider_height*win_height/2), pos=(0, 0), 
    labels=slider_labels, ticks=slider_ticks, granularity=slider_granularity, 
    style=['rating'], color=sliderColour, flip=False, depth=-3)
text_slider = visual.TextStim(win=win, name='text_slider', text=' ', 
    pos=(0, -.45), height=0.06, wrapWidth=None, ori=0, color='white', depth=-4.0)
text_data = visual.TextStim(win=win, name='text_data', text=' ', 
    pos=(0, -.45), height=0.03, wrapWidth=None, ori=0, color='white', depth=-5.0)
Idx = 0
for label in slider.labelObjs:
    label.height = .03  # Or whatever size you want
    label.wrapWidth = .2
    label.text=slider_labels[Idx]
    Idx+=1
slider.marker.color=marker_colour_discrete

#-----------------------
# Stimulus Presentation
#-----------------------

class Clock:
    def __init__(self):
        self.start_time = time.time()
    def getTime(self):
        return time.time() - self.start_time

def play_sound(file, volume):
	mySound = sound.Sound(file, preBuffer=-1, volume=volume)	
	mySound.play()
	core.wait(mySound.getDuration()+1)
     
def play_sound_rep(file, volume, n_repeat, gap, wait_time):
    time.sleep(wait_time)
    for r in range(n_repeat):
        mySound = sound.Sound(file, volume=volume)	
        repClock = Clock()
        cur_sound_start = repClock.getTime()
        mySound.play()
        while True:
            if repClock.getTime() > cur_sound_start + mySound.getDuration() + gap:
                break
            time.sleep(0.1)

if test_mode:
    n_stimulus_for_loop = test_n_stimulus
else :
    n_stimulus_for_loop = n_stimulus

break_count = 0

n_stimulus_each_break = math.ceil( n_stimulus_for_loop / (n_break+1) )

for sound_idx in range(n_stimulus_for_loop):
    cur_sound_name = sound_list['sound'][sound_idx]
    cur_type = sound_list['type'][sound_idx]
    cur_sound_path = os.path.join(thisDir, folder_input , cur_type, cur_sound_name)
    mySound = sound.Sound(cur_sound_path, preBuffer=-1, volume=volume)
    # get sound duration
    sound_list['duration'][sound_idx] = mySound.duration
    write_value_to_excel(os.path.join(thisExp.dataFileName+'.xlsx'),'SoundList',mySound.duration,cur_sound_name,'duration')

    #------------------------------
    # Block start
    #------------------------------
    if sound_idx == (break_count) * n_stimulus_each_break:
        text = visual.TextStim(win=win, text=" The No." + str(break_count+1) +" block starts soon."\
            "\n\n " + control_text + " to continue!",\
        color="white",
        contrast=contrast,\
        wrapWidth=max(win.size)/ratio_text_width,
        height= max(win.size)/ratio_text_size)
    
        text.autoDraw = True 
        win.flip()
        if control_mode == 'key':
            wait_for_enter_key(end_key_name)
        elif control_mode == 'mouse':
            wait_for_mouse_click(mouse)
        text.autoDraw = False
    
    text = visual.TextStim(win=win, text="Excerpt No. " + str(sound_idx+1) + " is about to start."\
                           	"\n\n" + control_text + " to continue!",\
        color="white",
        contrast=contrast,
        wrapWidth=max(win.size)/ratio_text_width,
        height= max(win.size)/ratio_text_size
        )
    text.autoDraw = True 
    win.flip()
    if control_mode == 'key':
        wait_for_enter_key(end_key_name)
    elif control_mode == 'mouse':
        wait_for_mouse_click(mouse)
    text.autoDraw = False    
    win.flip()  
    #-----------------------
    # pre-trial
    #-----------------------
    # show fixation
    fixation.autoDraw = True
    win.mouseVisible = False
    win.flip()
    # an interval before presentation

    core.wait(interval_cross_sound)
    # play sound
    play_sound(cur_sound_path, volume)
    fixation.autoDraw = False 
    win.flip()
    # rate the sound
    if control_mode == 'mouse':
        win.mouseVisible = True
    #-----------------------
    # pre-trial Rating
    #-----------------------
    slider.marker.color = marker_colour_discrete
    above_text = visual.TextStim(win=win, text="Please rate the excerpt on the slider.",\
        color="white",pos=(0, 0.3*win.size[1]/2), contrast=contrast, wrapWidth=max(win.size)/ratio_text_width, height= max(win.size)/ratio_text_size)
    oldRating = -1
    if control_mode == 'key':
        _ , cur_pre_trial_slider_data = slider_routine_discrete_key(win, slider, slider_shape, text_slider, text_data, 
                slider_orientation, slider_ticks, slider_granularity, slider_shape_width,
                   slider_decimals, slideSpeed, oldRating, above_text, step_size, end_key_name)
    elif control_mode =='mouse':
        _ , cur_pre_trial_slider_data = slider_routine_discrete(win, slider, slider_shape, text_slider, text_data, slider_orientation, 
                                                                    slider_ticks, slider_granularity, slider_shape_width, slider_decimals, sliding, 
                                                                    slideSpeed, oldRating, mouse, above_text)
    
    ratings = [item[0] for item in cur_pre_trial_slider_data]
    ratings = [cur_sound_name] + [cur_type] + ['pre_trial'] + ['rating'] + ratings
    # get pre-trial rating and save
    try:
        cur_pre_trial_rating = cur_pre_trial_slider_data[-1][0]
        oldRating = cur_pre_trial_rating
    except IndexError:
        cur_pre_trial_rating = error_slider_data_null
        oldRating = -1

    sound_list['pre-rating'][sound_idx] = cur_pre_trial_rating
    write_value_to_excel(os.path.join(thisExp.dataFileName+'.xlsx'),'SoundList',cur_pre_trial_rating,cur_sound_name,'pre-rating')
    times = [item[1] for item in cur_pre_trial_slider_data]
    times = [cur_sound_name] + [cur_type] + ['pre_trial'] + ['time'] + times
    write_vector_to_excel(os.path.join(thisExp.dataFileName+'.xlsx'), 'slider', ratings)
    write_vector_to_excel(os.path.join(thisExp.dataFileName+'.xlsx'), 'slider', times)

    #-------------------------------
    # pre-trial Rating CONFIRMATION
    #-------------------------------
    slider.marker.color = marker_colour_discrete
    above_text = visual.TextStim(win=win, text="Please confirm your rating.",\
        color="white",pos=(0, 0.3*win.size[1]/2), contrast=contrast, wrapWidth=max(win.size)/ratio_text_width, height= max(win.size)/ratio_text_size)
    if control_mode == 'key':
        _ , cur_pre_trial_cfm_slider_data = slider_routine_discrete_key(win, slider, slider_shape, text_slider, text_data, 
                slider_orientation, slider_ticks, slider_granularity, slider_shape_width,
                   slider_decimals, slideSpeed, oldRating, above_text, step_size, end_key_name)
    elif control_mode == 'mouse':
        _ , cur_pre_trial_cfm_slider_data = slider_routine_discrete(win, slider, slider_shape, text_slider, text_data, slider_orientation, 
                                                                    slider_ticks, slider_granularity, slider_shape_width, slider_decimals, sliding, 
                                                                    slideSpeed, oldRating, mouse, above_text)

    ratings = [item[0] for item in cur_pre_trial_cfm_slider_data]
    ratings = [cur_sound_name] + [cur_type] + ['pre_trial_confirm'] + ['rating'] + ratings
    # get pre-trial rating and save
    try:
        cur_pre_trial_cfm_rating = cur_pre_trial_cfm_slider_data[-1][0]
    except IndexError:
        cur_pre_trial_cfm_rating = error_slider_data_null
        
    write_value_to_excel(os.path.join(thisExp.dataFileName+'.xlsx'),'SoundList',cur_pre_trial_cfm_rating,cur_sound_name,'pre-rating_confirm')
    sound_list['pre-rating_confirm'][sound_idx] = cur_pre_trial_cfm_rating
    times = [item[1] for item in cur_pre_trial_cfm_slider_data]
    times = [cur_sound_name] + [cur_type] + ['pre_trial_confirm'] + ['time'] + times
    write_vector_to_excel(os.path.join(thisExp.dataFileName+'.xlsx'), 'slider', ratings)
    write_vector_to_excel(os.path.join(thisExp.dataFileName+'.xlsx'), 'slider', times)

    #----------------------------
    # trial (stimulus repetition)
    #----------------------------
    text = visual.TextStim(win=win, text="This excerpt will be repeated "+ str(n_repeat) + " times,"\
        "\n please rate it as repetition by moving mouse along the slider."\
        "\n\n The repetition will start soon."\
        "\n " + control_text + " to continue!",\
        color="white",
        contrast=contrast,\
        wrapWidth=max(win.size)/ratio_text_width,
        height= max(win.size)/ratio_text_size)
    text.autoDraw = True 
    win.flip()
    if control_mode == 'key':
        wait_for_enter_key(end_key_name)
    elif control_mode == 'mouse':
        wait_for_mouse_click(mouse)
    text.autoDraw = False    
    win.flip()  

    # set start position of mouse and marker
    initial_marker_pos = cur_pre_trial_cfm_rating if not cur_pre_trial_cfm_rating == error_slider_data_null else (slider_ticks[0] + slider_ticks[-1])/2
    oldRating = initial_marker_pos
    initial_rel_mouse_pos = [((initial_marker_pos-(slider_ticks[0] + slider_ticks[-1])/2) / (slider_ticks[-1] - slider_ticks[0])*2),0]
    initial_mouse_pos = [0,0]
    initial_mouse_pos[slider_orientation] = initial_rel_mouse_pos[slider_orientation] * slider_shape_width /2 * win.size[slider_orientation]

    slider.marker.color = marker_colour_continuous
    if control_mode == 'key':
        cur_trial_slider, cur_trial_slider_data  = slider_routine_continuous_key(win, slider, slider_shape, text_slider, text_data, 
            slider_orientation, slider_ticks, slider_granularity, slider_shape_width, slider_decimals, sliding, slideSpeed, oldRating,
            cur_sound_path, n_repeat, record_after, rep_interval_cross_sound, gap_time, volume, step_size)
    elif control_mode == 'mouse':
        cur_trial_slider, cur_trial_slider_data  = slider_routine_continuous(win, slider, slider_shape, text_slider, text_data, 
                slider_orientation, slider_ticks, slider_granularity, slider_shape_width, slider_decimals, sliding, slideSpeed, oldRating, mouse,
                cur_sound_path, n_repeat, record_after, rep_interval_cross_sound, gap_time, initial_mouse_pos, initial_marker_pos, volume)
        
    ratings = [item[0] for item in cur_trial_slider_data if item[1] >= rep_interval_cross_sound*1000]
    ratings = [cur_sound_name] + [cur_type] + ['trial'] + ['rating'] + ratings
    times = [item[1] - rep_interval_cross_sound*1000 for item in cur_trial_slider_data if item[1] >= rep_interval_cross_sound*1000]
    times = [cur_sound_name] + [cur_type] + ['trial'] + ['time'] + times
    write_vector_to_excel(os.path.join(thisExp.dataFileName+'.xlsx'), 'slider', ratings)
    write_vector_to_excel(os.path.join(thisExp.dataFileName+'.xlsx'), 'slider', times)    
    
    #-----------------------
    # post-trial
    #-----------------------
    text = visual.TextStim(win=win, text="This excerpt will be repeated one more times,"\
        "\n please look at the upcoming cross."\
        "\n\n " + control_text + " to continue!",\
        color="white",
        contrast=contrast,\
        wrapWidth=max(win.size)/ratio_text_width,
        height= max(win.size)/ratio_text_size)
    
    text.autoDraw = True 
    win.flip()
    if control_mode == 'key':
        wait_for_enter_key(end_key_name)
    elif control_mode == 'mouse':
        wait_for_mouse_click(mouse)
    text.autoDraw = False
    # show fixation
    fixation.autoDraw = True
    win.mouseVisible = False
    win.flip()
    # an interval before presentation
    core.wait(interval_cross_sound)
    # play sound
    play_sound(cur_sound_path, volume)
    fixation.autoDraw = False 
    win.flip()
    # rate the sound
    if control_mode == 'mouse':
        win.mouseVisible = True
    #-----------------------
    # post-trial Rating
    #-----------------------
    oldRating = -1
    slider.marker.color = marker_colour_discrete
    above_text = visual.TextStim(win=win, text="Please rate the excerpt on the slider.",\
        color="white",pos=(0, 0.3*win.size[1]/2), contrast=contrast, wrapWidth=max(win.size)/ratio_text_width, height= max(win.size)/ratio_text_size)
    if control_mode == 'key':
        _ , cur_post_trial_slider_data = slider_routine_discrete_key(win, slider, slider_shape, text_slider, text_data, 
                slider_orientation, slider_ticks, slider_granularity, slider_shape_width,
                   slider_decimals, slideSpeed, oldRating, above_text, step_size, end_key_name)
    elif control_mode == 'mouse':
        _ , cur_post_trial_slider_data = slider_routine_discrete(win, slider, slider_shape, text_slider, text_data, slider_orientation, 
                                                                    slider_ticks, slider_granularity, slider_shape_width, slider_decimals, sliding, 
                                                                    slideSpeed, oldRating, mouse, above_text)
    ratings = [item[0] for item in cur_post_trial_slider_data]
    ratings = [cur_sound_name] + [cur_type] + ['post_trial'] + ['rating'] + ratings
    # get post-trial rating and save
    try:
        cur_post_trial_rating = cur_post_trial_slider_data[-1][0]
    except IndexError:
        cur_post_trial_rating = error_slider_data_null
        
    sound_list['post-rating'][sound_idx] = cur_post_trial_rating
    write_value_to_excel(os.path.join(thisExp.dataFileName+'.xlsx'),'SoundList',cur_post_trial_rating,cur_sound_name,'post-rating')

    try:
        cur_post_trial_rating = cur_post_trial_slider_data[-1][0]
        oldRating = cur_post_trial_rating
    except IndexError:
        cur_post_trial_rating = error_slider_data_null
        oldRating = -1
    
    times = [item[1] for item in cur_post_trial_slider_data]
    times = [cur_sound_name] + [cur_type] + ['post_trial'] + ['time'] + times
    write_vector_to_excel(os.path.join(thisExp.dataFileName+'.xlsx'), 'slider', ratings)
    write_vector_to_excel(os.path.join(thisExp.dataFileName+'.xlsx'), 'slider', times)  

    #-------------------------------
    # post-trial Rating CONFIRMATION
    #-------------------------------
    slider.marker.color = marker_colour_discrete
    above_text = visual.TextStim(win=win, text="Please confirm your rating.",\
        color="white",pos=(0, 0.3*win.size[1]/2), contrast=contrast, wrapWidth=max(win.size)/ratio_text_width, height= max(win.size)/ratio_text_size)
    if control_mode == 'key':
        _ , cur_post_trial_cfm_slider_data = slider_routine_discrete_key(win, slider, slider_shape, text_slider, text_data, 
                slider_orientation, slider_ticks, slider_granularity, slider_shape_width,
                   slider_decimals, slideSpeed, oldRating, above_text, step_size, end_key_name)
    elif control_mode == 'mouse':
        _ , cur_post_trial_cfm_slider_data = slider_routine_discrete(win, slider, slider_shape, text_slider, text_data, slider_orientation, 
                                                                    slider_ticks, slider_granularity, slider_shape_width, slider_decimals, sliding, 
                                                                    slideSpeed, oldRating, mouse, above_text)
    ratings = [item[0] for item in cur_post_trial_cfm_slider_data]
    ratings = [cur_sound_name] + [cur_type] + ['post_trial_confirm'] + ['rating'] + ratings
    # get post-trial rating and save
    try:
        cur_post_trial_cfm_rating = cur_post_trial_cfm_slider_data[-1][0]
    except IndexError:
        cur_post_trial_cfm_rating = error_slider_data_null

    write_value_to_excel(os.path.join(thisExp.dataFileName+'.xlsx'),'SoundList',cur_post_trial_cfm_rating,cur_sound_name,'post-rating_confirm')
    sound_list['post-rating_confirm'][sound_idx] = cur_post_trial_cfm_rating
    times = [item[1] for item in cur_post_trial_cfm_slider_data]
    times = [cur_sound_name] + [cur_type] + ['post_trial_confirm'] + ['time'] + times
    write_vector_to_excel(os.path.join(thisExp.dataFileName+'.xlsx'), 'slider', ratings)
    write_vector_to_excel(os.path.join(thisExp.dataFileName+'.xlsx'), 'slider', times)

    #------------------------------
    # Break Time
    #------------------------------
    if sound_idx == (break_count+1) * n_stimulus_each_break - 1 :
        text = visual.TextStim(win=win, text=" Well done! You finished the No." + str(break_count+1) +" block. It's break time now."\
            "\n If you'd like to have a break outside the room, please inform the experimenter."\
            "\n\n " + control_text + " to continue!",\
        color="white",
        contrast=contrast,\
        wrapWidth=max(win.size)/ratio_text_width,
        height= max(win.size)/ratio_text_size)
    
        text.autoDraw = True 
        win.flip()
        if control_mode == 'key':
            wait_for_enter_key(end_key_name)
        elif control_mode == 'mouse':
            wait_for_mouse_click(mouse)
        text.autoDraw = False
        break_count += 1

     

#-----------------------
# Ending
#-----------------------
text = visual.TextStim(win=win, text="Congratulations! You finished all excerpts!"\
                       "\n\n Please contact the experimenter."\
                       "\n\n Thanks for joining our experiment! Wish you a good day :)",\
    color="white",
    contrast=contrast,\
    wrapWidth=max(win.size)/ratio_text_width,
    height= max(win.size)/ratio_text_size)
text.autoDraw = True 
win.flip()
if control_mode == 'key':
    wait_for_enter_key(end_key_name)
elif control_mode == 'mouse':
    wait_for_mouse_click(mouse)
text.autoDraw = False
win.close()

saveData(thisExp=thisExp)

#quit(thisExp=thisExp, win=win)